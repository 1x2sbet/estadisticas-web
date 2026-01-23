from playwright.sync_api import sync_playwright
import pandas as pd
import time
import os
import re
from openpyxl import load_workbook

# -----------------------
# CONFIGURACIÓN
# -----------------------
EXCEL_LINKS = r"C:\Users\luisf\OneDrive\Escritorio\INFO LUIS\JUEGOS\DATOS PYTHON\DATOS EXTRAIDOS PYTHON.xlsx"
HOJA_LINKS = "LISTADOLINKS"
OUTPUT_FILE = r"C:\Users\luisf\OneDrive\Escritorio\INFO LUIS\JUEGOS\DATOS PYTHON\DATOS EXTRAIDOS PYTHON.xlsx"

# -----------------------
# LEER LINKS DESDE EXCEL
# -----------------------
def leer_links_desde_excel():
    try:
        df = pd.read_excel(EXCEL_LINKS, sheet_name=HOJA_LINKS)
        df.columns = [c.strip().upper() for c in df.columns]

        if "NP BETPLAY" not in df.columns:
            df["NP BETPLAY"] = ""

        if not {"ENCENDIDO", "LIGA", "BETPLAY"}.issubset(df.columns):
            print("❌ El archivo debe tener las columnas: 'ENCENDIDO', 'LIGA' y 'BETPLAY'.")
            return df, []

        df_activas = df[df["ENCENDIDO"].astype(str).str.upper().str.strip() == "ACTIVO"]
        df_activas = df_activas[df_activas["BETPLAY"].astype(str).str.startswith("http")]

        urls = list(zip(df_activas["LIGA"], df_activas["BETPLAY"]))
        print(f"📚 {len(urls)} ligas activas para procesar.")
        return df, urls

    except Exception as e:
        print(f"❌ Error leyendo Excel: {e}")
        return pd.DataFrame(), []

# -----------------------
# EXTRAER PARTIDOS
# -----------------------
def extraer_partidos(page, liga, url, recargar=False):
    try:
        page.goto(url, timeout=60000)
        page.wait_for_timeout(9000)

        tiene_section = page.query_selector("main.KambiBC-sports-hub section")
        if not tiene_section:
            print(f"⚠️ Liga sin partidos: {liga}")
            return 0, []

        cuerpo_texto = page.inner_text("body") if page.query_selector("body") else ""
        if "Posición final" in cuerpo_texto:
            try:
                fecha_span = page.query_selector("span.KambiBC-event-item__start-time--date")
                if fecha_span:
                    fecha_texto = fecha_span.inner_text().strip()
                    print(f"⚠️ Liga sin iniciar: {liga} inicia el {fecha_texto}")
                else:
                    html = page.content()
                    m = re.search(r"([0-9]{1,2}\s+de\s+[a-záéíóúñ]+(?:\s+[a-z]+)?\s+de\s+[0-9]{4})", html, flags=re.IGNORECASE)
                    if m:
                        fecha_texto = m.group(1)
                        print(f"⚠️ Liga sin iniciar: {liga} inicia el {fecha_texto}")
                    else:
                        print(f"⚠️ Liga sin iniciar: {liga} (fecha no detectada)")
            except:
                print(f"⚠️ Liga sin iniciar: {liga} (fecha no detectada)")
            return "NO INICIADO", []

        partidos = []
        items = page.query_selector_all("li.KambiBC-sandwich-filter__event-list-item")

        for p in items:
            try:
                equipos = p.query_selector_all("div.KambiBC-event-participants__name-participant-name")
                local = equipos[0].inner_text().strip() if len(equipos) > 0 else ""
                visitante = equipos[1].inner_text().strip() if len(equipos) > 1 else ""

                fecha_span = p.query_selector("span.KambiBC-event-item__start-time--date")
                hora_span = p.query_selector("span.KambiBC-event-item__start-time--time")
                fecha = fecha_span.inner_text().strip() if fecha_span else ""
                hora = hora_span.inner_text().strip() if hora_span else ""
                fecha_hora = f"{fecha}, {hora}" if fecha or hora else ""

                cuotas = p.query_selector_all("div.KambiBC-bet-offer--onecrosstwo button.KambiBC-betty-outcome")
                c1 = cuotas[0].inner_text().strip() if len(cuotas) > 0 else ""
                cx = cuotas[1].inner_text().strip() if len(cuotas) > 1 else ""
                c2 = cuotas[2].inner_text().strip() if len(cuotas) > 2 else ""

                partidos.append({
                    "Liga": liga,
                    "Fecha": fecha_hora,
                    "Local": local,
                    "Visitante": visitante,
                    "Cuota Local": c1,
                    "Cuota Empate": cx,
                    "Cuota Visitante": c2
                })
            except:
                continue

        if not partidos and not recargar:
            print(f"🔄 Sin datos en {liga}, recargando página y esperando 20s...")
            page.reload()
            page.wait_for_timeout(20000)
            return extraer_partidos(page, liga, url, recargar=True)

        if partidos:
            print(f"✅ {len(partidos)} partidos extraídos de {liga}")
            return len(partidos), partidos
        else:
            print(f"⚠️ No se encontraron partidos en {liga} tras recarga")
            return 0, []

    except Exception as e:
        print(f"❌ Error en {liga}: {e}")
        return 0, []

# -----------------------
# PROGRAMA PRINCIPAL
# -----------------------
start_time = time.time()
df_links, urls = leer_links_desde_excel()
total_partidos_extraidos = 0
np_por_liga = {}

if not urls:
    print("⚠️ No hay ligas activas para procesar.")
else:
    with sync_playwright() as p:
        browser = p.chromium.launch_persistent_context(
            user_data_dir=r"C:\BetplayProfile",
            headless=False,
            args=["--start-maximized"]
        )
        all_results = []

        for liga, url in urls:
            page = browser.new_page()
            cantidad, resultados = extraer_partidos(page, liga, url)
            all_results.extend(resultados)
            np_por_liga[liga] = cantidad
            if isinstance(cantidad, int):
                total_partidos_extraidos += cantidad
            try:
                page.close()
            except:
                pass
            time.sleep(2)

        try:
            browser.close()
        except:
            pass

    df = pd.DataFrame(all_results)
    if not df.empty:
        if os.path.exists(OUTPUT_FILE):
            with pd.ExcelWriter(OUTPUT_FILE, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
                df.to_excel(writer, sheet_name="BETPLAY", index=False)
        else:
            with pd.ExcelWriter(OUTPUT_FILE, engine="openpyxl") as writer:
                df.to_excel(writer, sheet_name="BETPLAY", index=False)

        print(f"📊 Datos actualizados en la hoja 'BETPLAY' del archivo:\n   {OUTPUT_FILE}")
    else:
        print("⚠️ No se extrajo ningún partido.")

    try:
        wb = load_workbook(EXCEL_LINKS)
        if HOJA_LINKS not in wb.sheetnames:
            print(f"❌ No se encontró la hoja '{HOJA_LINKS}' en el archivo Excel.")
        else:
            ws = wb[HOJA_LINKS]

            headers = [cell.value for cell in ws[1]]
            if "LIGA" not in headers or "BETPLAY" not in headers or "ENCENDIDO" not in headers:
                print("❌ La hoja no contiene las columnas necesarias ('LIGA', 'ENCENDIDO', 'BETPLAY').")
            else:
                idx_liga = headers.index("LIGA") + 1
                idx_encendido = headers.index("ENCENDIDO") + 1
                idx_betplay = headers.index("BETPLAY") + 1
                if "NP BETPLAY" not in headers:
                    ws.cell(row=1, column=len(headers) + 1).value = "NP BETPLAY"
                    idx_np = len(headers) + 1
                else:
                    idx_np = headers.index("NP BETPLAY") + 1

                for row in range(2, ws.max_row + 1):
                    liga = str(ws.cell(row=row, column=idx_liga).value or "").strip()
                    encendido = str(ws.cell(row=row, column=idx_encendido).value or "").strip().upper()
                    link = str(ws.cell(row=row, column=idx_betplay).value or "").strip()

                    valor_np = ""
                    if link == "" or not link.startswith("http"):
                        valor_np = ""
                    elif encendido != "ACTIVO":
                        valor_np = ""
                    elif liga in np_por_liga:
                        valor = np_por_liga[liga]
                        if valor == "NO INICIADO":
                            valor_np = "NO INICIADO"
                        elif isinstance(valor, int) and valor == 0:
                            valor_np = 0
                        else:
                            valor_np = valor

                    ws.cell(row=row, column=idx_np, value=valor_np)

            wb.save(EXCEL_LINKS)
            wb.close()
            print(f"💾 Actualizado solo la columna 'NP BETPLAY' en '{HOJA_LINKS}'.")
    except Exception as e:
        print(f"❌ Error guardando NP BETPLAY en {EXCEL_LINKS}: {e}")

    print(f"\n⏱️ Tiempo total de ejecución: {time.time() - start_time:.2f} segundos")
    print(f"📊 Total de partidos extraídos: {total_partidos_extraidos}")

